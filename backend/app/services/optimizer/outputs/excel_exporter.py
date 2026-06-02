"""Excel 导出器 — 导出4个Sheet"""
import io
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from app.services.optimizer.models.task_model import Task
from app.services.optimizer.models.solution import Solution
from app.services.optimizer.outputs.table_builder import (
    build_summary_stats,
    build_route_summary_table,
    build_village_detail_table,
    build_drone_detail_table,
)


def export_excel(task: Task, solution: Solution) -> bytes:
    """
    导出 Excel 文件（4个Sheet）。

    Sheet 1: 路径汇总
    Sheet 2: 村庄配送详情
    Sheet 3: 无人机配送详情
    Sheet 4: 任务执行时间轴

    返回: Excel 文件的 bytes
    """
    wb = Workbook()

    # 样式
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def style_header(ws, row=1):
        for cell in ws[row]:
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

    def style_data(ws, start_row=2):
        for row in ws.iter_rows(min_row=start_row):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center")

    # ── Sheet 1: 路径汇总 ──
    ws1 = wb.active
    ws1.title = "路径汇总"
    route_table = build_route_summary_table(solution, task)
    ws1.append(["路径编号", "无人机", "路径", "途径节点", "距离(km)", "时间(min)", "配送重量(kg)", "目标村庄", "趟次"])
    for r in route_table:
        ws1.append([
            r["route_id"], r["drone_name"], r["route_path"], r["route_nodes"],
            r["distance"], r["time"], r["weight"], r["village_name"], r["trip_count"],
        ])
    # 总计行
    ws1.append([
        "总计", "-", "-", "-",
        round(solution.total_distance, 2),
        round(solution.total_time, 2),
        round(solution.total_delivered, 2),
        "-",
        solution.total_trips,
    ])
    style_header(ws1)
    style_data(ws1)

    # ── Sheet 2: 村庄配送详情 ──
    ws2 = wb.create_sheet("村庄配送详情")
    village_table = build_village_detail_table(solution, task)
    ws2.append(["村庄编号", "村庄名称", "需求重量(kg)", "配送无人机", "无人机配送重量(kg)", "趟次", "单程距离(km)", "特殊要求"])
    for v in village_table:
        ws2.append([
            v["village_id"], v["village_name"], v["demand_weight"],
            v["drone_name"], v["drone_weight"], v["trip_count"],
            v["one_way_distance"], v["special_req"],
        ])
    style_header(ws2)
    style_data(ws2)

    # ── Sheet 3: 无人机配送详情 ──
    ws3 = wb.create_sheet("无人机配送详情")
    drone_table = build_drone_detail_table(solution, task)
    ws3.append(["无人机编号", "机型", "速度(m/s)", "最大载重(kg)", "总飞行距离(km)", "总飞行时间(min)", "总趟次", "服务村庄", "备注"])
    for d in drone_table:
        ws3.append([
            d["drone_id"], d["drone_type"], d["speed_ms"], d["max_payload"],
            d["total_distance"], d["total_time"], d["total_trips"],
            d["villages"], d["note"],
        ])
    style_header(ws3)
    style_data(ws3)

    # ── Sheet 4: 任务执行时间轴 ──
    ws4 = wb.create_sheet("任务执行时间轴")
    ws4.append(["趟次ID", "无人机", "村庄", "开始时间(min)", "结束时间(min)", "飞行时间(min)", "距离(km)", "载重(kg)", "状态"])
    for trip in sorted(solution.trips, key=lambda t: t.start_time):
        ws4.append([
            trip.trip_id, trip.drone_type, trip.village_name,
            round(trip.start_time, 2), round(trip.end_time, 2),
            round(trip.total_time, 2), round(trip.total_distance, 2),
            round(trip.load, 2),
            "可行" if trip.feasible else "不可行",
        ])
    style_header(ws4)
    style_data(ws4)

    # 保存到 bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
